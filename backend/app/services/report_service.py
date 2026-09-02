import io
from datetime import datetime
from typing import Any, Dict, List, Optional
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape, portrait
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from app.core.config import settings
from app.core.database import db
from app.services.nasabah_service import nasabah_service
from app.utils.currency import format_rupiah
from app.utils.formatting import format_date, format_datetime, format_kg


class ReportService:
    # Styling helpers for Excel
    @staticmethod
    def _create_excel_header(ws, title: str, subtitle: str, col_span: int):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_span)
        ws.cell(row=1, column=1, value=settings.BANK_NAME.upper())
        ws.cell(row=1, column=1).font = Font(name="Arial", size=14, bold=True, color="15803D")
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_span)
        ws.cell(row=2, column=1, value=title)
        ws.cell(row=2, column=1).font = Font(name="Arial", size=11, bold=True, color="1F2937")
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")

        if subtitle:
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=col_span)
            ws.cell(row=3, column=1, value=subtitle)
            ws.cell(row=3, column=1).font = Font(name="Arial", size=9, italic=True, color="4B5563")
            ws.cell(row=3, column=1).alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 18
        if subtitle:
            ws.row_dimensions[3].height = 16

    @staticmethod
    def _auto_fit_columns(ws, min_col=1, max_col=None):
        max_col = max_col or ws.max_column
        for col in range(min_col, max_col + 1):
            col_letter = get_column_letter(col)
            max_len = 0
            for cell in ws[col_letter]:
                if cell.row > 4 and cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    @staticmethod
    def _get_pdf_styles():
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "PDFTitle",
            parent=styles["Heading2"],
            fontSize=13,
            leading=16,
            textColor=colors.HexColor("#15803d"),
            alignment=1,
            spaceAfter=2,
        )
        sub_style = ParagraphStyle(
            "PDFSub",
            parent=styles["Normal"],
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#4b5563"),
            alignment=1,
            spaceAfter=6,
        )
        
        # Table content cell styles (with auto text-wrapping via Paragraph)
        cell_left = ParagraphStyle("CL", parent=styles["Normal"], fontSize=8, leading=10, textColor=colors.HexColor("#1f2937"), alignment=0)
        cell_center = ParagraphStyle("CC", parent=styles["Normal"], fontSize=8, leading=10, textColor=colors.HexColor("#1f2937"), alignment=1)
        cell_right = ParagraphStyle("CR", parent=styles["Normal"], fontSize=8, leading=10, textColor=colors.HexColor("#1f2937"), alignment=2)
        
        # Bold cell styles
        cell_bold_left = ParagraphStyle("CBL", parent=cell_left, fontName="Helvetica-Bold")
        cell_bold_center = ParagraphStyle("CBC", parent=cell_center, fontName="Helvetica-Bold")
        cell_bold_right = ParagraphStyle("CBR", parent=cell_right, fontName="Helvetica-Bold")
        
        # Table Header styles
        th_left = ParagraphStyle("THL", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=colors.HexColor("#166534"), alignment=0)
        th_center = ParagraphStyle("THC", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=colors.HexColor("#166534"), alignment=1)
        th_right = ParagraphStyle("THR", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=colors.HexColor("#166534"), alignment=2)
        
        return {
            "title": title_style,
            "sub": sub_style,
            "cell_left": cell_left,
            "cell_center": cell_center,
            "cell_right": cell_right,
            "cell_bold_left": cell_bold_left,
            "cell_bold_center": cell_bold_center,
            "cell_bold_right": cell_bold_right,
            "th_left": th_left,
            "th_center": th_center,
            "th_right": th_right,
        }

    # 1. TRANSACTIONS REPORT
    @classmethod
    def get_transactions_report_data(
        cls,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        type_filter: Optional[str] = None,
        category_id: Optional[str] = None,
        nasabah_id: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        query = """
            SELECT t.*, 
                   n.customer_id as nasabah_customer_id, 
                   n.name as nasabah_name, 
                   n.nik as nasabah_nik,
                   c.name as category_name
            FROM transactions t
            JOIN nasabah n ON t.nasabah_id = n.id
            LEFT JOIN waste_categories c ON t.category_id = c.id
            WHERE 1=1
        """
        params: List[Any] = []
        if start_date:
            query += " AND date(t.transaction_date) >= date(?)"
            params.append(start_date)
        if end_date:
            query += " AND date(t.transaction_date) <= date(?)"
            params.append(end_date)
        if type_filter:
            query += " AND t.type = ?"
            params.append(type_filter)
        if category_id:
            query += " AND t.category_id = ?"
            params.append(category_id)
        if nasabah_id:
            query += " AND t.nasabah_id = ?"
            params.append(nasabah_id)

        query += " ORDER BY t.transaction_date ASC, t.created_at ASC"
        return db.fetch_all(query, params)

    @classmethod
    def export_transactions_excel(
        cls, start_date: Optional[str] = None, end_date: Optional[str] = None, **kwargs
    ) -> bytes:
        data = cls.get_transactions_report_data(start_date=start_date, end_date=end_date, **kwargs)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Laporan Transaksi"

        subtitle = f"Periode: {format_date(start_date) if start_date else 'Awal'} s/d {format_date(end_date) if end_date else 'Sekarang'}"
        cls._create_excel_header(ws, "LAPORAN TRANSAKSI TABUNGAN BANK SAMPAH", subtitle, 12)

        headers = [
            "No", "No. Transaksi", "Tanggal", "ID Nasabah", "Nama Nasabah",
            "Jenis", "Kategori", "Berat (kg)", "Harga/kg", "Debit (Rp)", "Kredit (Rp)", "Saldo Akhir (Rp)"
        ]

        header_row = 5
        header_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        header_font = Font(name="Arial", size=9, bold=True, color="166534")
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'), right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'), bottom=Side(style='thin', color='E5E7EB')
        )

        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        ws.row_dimensions[header_row].height = 22

        current_row = 6
        total_debit = 0
        total_credit = 0

        for i, item in enumerate(data, 1):
            weight_kg = (item["weight_gram"] / 1000.0) if item.get("weight_gram") else 0
            price_per_kg = item.get("price_per_kg") or 0
            debit = item.get("debit") or 0
            credit = item.get("credit") or 0
            total_debit += debit
            total_credit += credit

            row_values = [
                i,
                item["transaction_no"],
                format_datetime(item["transaction_date"]),
                item["nasabah_customer_id"],
                item["nasabah_name"],
                item["type"],
                item.get("category_name") or "-",
                weight_kg if weight_kg > 0 else "-",
                price_per_kg if price_per_kg > 0 else "-",
                debit,
                credit,
                item["balance_after"],
            ]

            for col_num, val in enumerate(row_values, 1):
                cell = ws.cell(row=current_row, column=col_num, value=val)
                cell.font = Font(name="Arial", size=9)
                cell.border = thin_border
                if col_num in (8, 9, 10, 11, 12):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(val, (int, float)) and col_num >= 9:
                        cell.number_format = "#,##0"
                elif col_num in (1, 3, 6):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1

        # Summary row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
        summary_cell = ws.cell(row=current_row, column=1, value="TOTAL")
        summary_cell.font = Font(name="Arial", size=9, bold=True)
        summary_cell.alignment = Alignment(horizontal="center", vertical="center")
        summary_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

        for col_num in range(1, 13):
            cell = ws.cell(row=current_row, column=col_num)
            cell.fill = summary_fill
            cell.border = thin_border
            cell.font = Font(name="Arial", size=9, bold=True)

        debit_sum_cell = ws.cell(row=current_row, column=10, value=total_debit)
        debit_sum_cell.number_format = "#,##0"
        debit_sum_cell.alignment = Alignment(horizontal="right", vertical="center")

        credit_sum_cell = ws.cell(row=current_row, column=11, value=total_credit)
        credit_sum_cell.number_format = "#,##0"
        credit_sum_cell.alignment = Alignment(horizontal="right", vertical="center")

        cls._auto_fit_columns(ws, 1, 12)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @classmethod
    def export_transactions_pdf(
        cls, start_date: Optional[str] = None, end_date: Optional[str] = None, **kwargs
    ) -> bytes:
        data = cls.get_transactions_report_data(start_date=start_date, end_date=end_date, **kwargs)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20
        )

        st = cls._get_pdf_styles()

        elements = [
            Paragraph(settings.BANK_NAME.upper(), st["title"]),
            Paragraph("LAPORAN TRANSAKSI TABUNGAN BANK SAMPAH", st["title"]),
            Paragraph(f"Periode: {format_date(start_date) if start_date else 'Awal'} s/d {format_date(end_date) if end_date else 'Sekarang'}", st["sub"]),
            Spacer(1, 10),
            HRFlowable(width="100%", thickness=1, color=colors.HexColor("#d1d5db"), spaceAfter=10)
        ]

        table_data = [[
            Paragraph("<b>No</b>", st["th_center"]),
            Paragraph("<b>No. Transaksi</b>", st["th_center"]),
            Paragraph("<b>Tanggal</b>", st["th_center"]),
            Paragraph("<b>ID Nasabah</b>", st["th_center"]),
            Paragraph("<b>Nama Nasabah</b>", st["th_left"]),
            Paragraph("<b>Jenis</b>", st["th_center"]),
            Paragraph("<b>Kelompok Sampah</b>", st["th_left"]),
            Paragraph("<b>Berat (kg)</b>", st["th_right"]),
            Paragraph("<b>Debit (Rp)</b>", st["th_right"]),
            Paragraph("<b>Kredit (Rp)</b>", st["th_right"]),
            Paragraph("<b>Saldo (Rp)</b>", st["th_right"])
        ]]

        tot_debit = 0
        tot_credit = 0

        for i, item in enumerate(data, 1):
            w = (item["weight_gram"] / 1000.0) if item.get("weight_gram") else 0
            d = item.get("debit") or 0
            c = item.get("credit") or 0
            tot_debit += d
            tot_credit += c

            table_data.append([
                Paragraph(str(i), st["cell_center"]),
                Paragraph(item["transaction_no"], st["cell_center"]),
                Paragraph(format_datetime(item["transaction_date"]), st["cell_center"]),
                Paragraph(item["nasabah_customer_id"], st["cell_center"]),
                Paragraph(item["nasabah_name"], st["cell_left"]),
                Paragraph(f"<b>{item['type']}</b>", st["cell_bold_center"]),
                Paragraph(item.get("category_name") or "-", st["cell_left"]),
                Paragraph(f"{w:.3f}" if w > 0 else "-", st["cell_right"]),
                Paragraph(format_rupiah(d) if d > 0 else "-", st["cell_right"]),
                Paragraph(format_rupiah(c) if c > 0 else "-", st["cell_right"]),
                Paragraph(format_rupiah(item["balance_after"]), st["cell_right"]),
            ])

        # Summary Row
        table_data.append([
            Paragraph("<b>TOTAL</b>", st["cell_bold_center"]),
            "", "", "", "", "", "", "",
            Paragraph(f"<b>{format_rupiah(tot_debit)}</b>", st["cell_bold_right"]),
            Paragraph(f"<b>{format_rupiah(tot_credit)}</b>", st["cell_bold_right"]),
            ""
        ])

        t = Table(table_data, colWidths=[25, 95, 75, 75, 115, 50, 105, 55, 70, 70, 70])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f0fdf4")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#166534")),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('SPAN', (0, -1), (7, -1)),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#f9fafb")),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(t)
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    # 2. CATEGORY RECAP REPORT
    @classmethod
    def get_category_recap_data(cls, start_date: Optional[str] = None, end_date: Optional[str] = None) -> List[Dict[str, Any]]:
        query = """
            SELECT c.id, c.name,
                   COUNT(t.id) as total_trx,
                   COALESCE(SUM(t.weight_gram), 0) as total_weight_gram,
                   COALESCE(SUM(t.credit), 0) as total_amount
            FROM waste_categories c
            LEFT JOIN transactions t ON c.id = t.category_id AND t.type = 'SETOR'
        """
        params: List[Any] = []
        if start_date:
            query += " AND date(t.transaction_date) >= date(?)"
            params.append(start_date)
        if end_date:
            query += " AND date(t.transaction_date) <= date(?)"
            params.append(end_date)

        query += " GROUP BY c.id, c.name ORDER BY total_amount DESC"
        return db.fetch_all(query, params)

    @classmethod
    def export_category_recap_excel(cls, start_date: Optional[str] = None, end_date: Optional[str] = None) -> bytes:
        data = cls.get_category_recap_data(start_date, end_date)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rekap Kategori"

        subtitle = f"Periode: {format_date(start_date) if start_date else 'Awal'} s/d {format_date(end_date) if end_date else 'Sekarang'}"
        cls._create_excel_header(ws, "REKAPITULASI SETORAN PER KATEGORI SAMPAH", subtitle, 5)

        headers = ["No", "Nama Kategori", "Jumlah Transaksi Setor", "Total Berat (kg)", "Total Nilai Setoran (Rp)"]
        header_row = 5
        header_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        header_font = Font(name="Arial", size=9, bold=True, color="166534")
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'), right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'), bottom=Side(style='thin', color='E5E7EB')
        )

        for col_num, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        current_row = 6
        tot_trx = 0
        tot_weight = 0
        tot_amt = 0

        for i, item in enumerate(data, 1):
            w_kg = item["total_weight_gram"] / 1000.0
            tot_trx += item["total_trx"]
            tot_weight += w_kg
            tot_amt += item["total_amount"]

            ws.cell(row=current_row, column=1, value=i).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=2, value=item["name"]).alignment = Alignment(horizontal="left")
            ws.cell(row=current_row, column=3, value=item["total_trx"]).alignment = Alignment(horizontal="right")
            ws.cell(row=current_row, column=4, value=w_kg).alignment = Alignment(horizontal="right")
            ws.cell(row=current_row, column=5, value=item["total_amount"]).alignment = Alignment(horizontal="right")
            ws.cell(row=current_row, column=5).number_format = "#,##0"

            for c in range(1, 6):
                ws.cell(row=current_row, column=c).border = thin_border
                ws.cell(row=current_row, column=c).font = Font(name="Arial", size=9)
            current_row += 1

        # Summary
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        ws.cell(row=current_row, column=1, value="TOTAL").alignment = Alignment(horizontal="center")
        ws.cell(row=current_row, column=3, value=tot_trx).alignment = Alignment(horizontal="right")
        ws.cell(row=current_row, column=4, value=tot_weight).alignment = Alignment(horizontal="right")
        ws.cell(row=current_row, column=5, value=tot_amt).alignment = Alignment(horizontal="right")
        ws.cell(row=current_row, column=5).number_format = "#,##0"

        for c in range(1, 6):
            cell = ws.cell(row=current_row, column=c)
            cell.border = thin_border
            cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            cell.font = Font(name="Arial", size=9, bold=True)

        cls._auto_fit_columns(ws, 1, 5)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @classmethod
    def export_category_recap_pdf(cls, start_date: Optional[str] = None, end_date: Optional[str] = None) -> bytes:
        data = cls.get_category_recap_data(start_date, end_date)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=portrait(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)

        st = cls._get_pdf_styles()

        elements = [
            Paragraph(settings.BANK_NAME.upper(), st["title"]),
            Paragraph("REKAPITULASI SETORAN PER KELOMPOK SAMPAH", st["title"]),
            Paragraph(f"Periode: {format_date(start_date) if start_date else 'Awal'} s/d {format_date(end_date) if end_date else 'Sekarang'}", st["sub"]),
            Spacer(1, 10),
            HRFlowable(width="100%", thickness=1, color=colors.HexColor("#d1d5db"), spaceAfter=10)
        ]

        table_data = [[
            Paragraph("<b>No</b>", st["th_center"]),
            Paragraph("<b>Kelompok Sampah</b>", st["th_left"]),
            Paragraph("<b>Jumlah Transaksi</b>", st["th_center"]),
            Paragraph("<b>Total Berat (kg)</b>", st["th_right"]),
            Paragraph("<b>Total Nilai (Rp)</b>", st["th_right"])
        ]]

        tot_trx = 0
        tot_weight = 0
        tot_amt = 0

        for i, item in enumerate(data, 1):
            w = item["total_weight_gram"] / 1000.0
            tot_trx += item["total_trx"]
            tot_weight += w
            tot_amt += item["total_amount"]

            table_data.append([
                Paragraph(str(i), st["cell_center"]),
                Paragraph(item["name"], st["cell_left"]),
                Paragraph(str(item["total_trx"]), st["cell_center"]),
                Paragraph(f"{w:.3f}", st["cell_right"]),
                Paragraph(format_rupiah(item["total_amount"]), st["cell_right"]),
            ])

        table_data.append([
            Paragraph("<b>TOTAL</b>", st["cell_bold_center"]),
            "",
            Paragraph(f"<b>{tot_trx}</b>", st["cell_bold_center"]),
            Paragraph(f"<b>{tot_weight:.3f}</b>", st["cell_bold_right"]),
            Paragraph(f"<b>{format_rupiah(tot_amt)}</b>", st["cell_bold_right"]),
        ])

        t = Table(table_data, colWidths=[30, 190, 95, 95, 135])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f0fdf4")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#166534")),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('SPAN', (0, -1), (1, -1)),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#f9fafb")),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(t)
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    # 3. NASABAH & BALANCE REPORT
    @classmethod
    def export_nasabah_excel(cls) -> bytes:
        nasabahs, _ = nasabah_service.list_nasabah(page=1, page_size=10000)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daftar Nasabah"

        cls._create_excel_header(ws, "LAPORAN DAFTAR NASABAH & SALDO TABUNGAN", f"Per Tanggal: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 15)

        headers = [
            "No", "ID Nasabah", "NIK", "Nama Nasabah", "Kategori", "No. HP", "Alamat",
            "RT", "RW", "Kelurahan", "Kecamatan", "Kabupaten/Kota",
            "Status", "Sumber", "Saldo Terakhir (Rp)"
        ]
        header_row = 5
        header_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        header_font = Font(name="Arial", size=9, bold=True, color="166534")
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'), right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'), bottom=Side(style='thin', color='E5E7EB')
        )

        for col_num, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        current_row = 6
        total_balance = 0

        for i, n in enumerate(nasabahs, 1):
            bal = n.get("balance", 0)
            total_balance += bal

            ws.cell(row=current_row, column=1, value=i).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=2, value=n["customer_id"]).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=3, value=f"'{n['nik']}").alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=4, value=n["name"]).alignment = Alignment(horizontal="left")
            ws.cell(row=current_row, column=5, value=n.get("nasabah_category") or "Rumah Tangga/Individu").alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=6, value=n["phone"]).alignment = Alignment(horizontal="left")
            ws.cell(row=current_row, column=7, value=n.get("address") or "-").alignment = Alignment(horizontal="left")
            ws.cell(row=current_row, column=8, value=n.get("rt") or "-").alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=9, value=n.get("rw") or "-").alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=10, value=n.get("kelurahan") or "-").alignment = Alignment(horizontal="left")
            ws.cell(row=current_row, column=11, value=n.get("kecamatan") or "-").alignment = Alignment(horizontal="left")
            ws.cell(row=current_row, column=12, value=n.get("kabupaten_kota") or "-").alignment = Alignment(horizontal="left")
            ws.cell(row=current_row, column=13, value=n["status"]).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=14, value=n["registration_source"]).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=15, value=bal).alignment = Alignment(horizontal="right")
            ws.cell(row=current_row, column=15).number_format = "#,##0"

            for c in range(1, 16):
                ws.cell(row=current_row, column=c).border = thin_border
                ws.cell(row=current_row, column=c).font = Font(name="Arial", size=9)
            current_row += 1

        # Total
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=14)
        ws.cell(row=current_row, column=1, value="TOTAL SALDO NASABAH").alignment = Alignment(horizontal="center")
        ws.cell(row=current_row, column=15, value=total_balance).alignment = Alignment(horizontal="right")
        ws.cell(row=current_row, column=15).number_format = "#,##0"

        for c in range(1, 16):
            cell = ws.cell(row=current_row, column=c)
            cell.border = thin_border
            cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            cell.font = Font(name="Arial", size=9, bold=True)

        cls._auto_fit_columns(ws, 1, 15)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @classmethod
    def export_nasabah_pdf(cls) -> bytes:
        nasabahs, _ = nasabah_service.list_nasabah(page=1, page_size=10000)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=portrait(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)

        st = cls._get_pdf_styles()

        elements = [
            Paragraph(settings.BANK_NAME.upper(), st["title"]),
            Paragraph("LAPORAN DAFTAR NASABAH & SALDO TABUNGAN", st["title"]),
            Paragraph(f"Per Tanggal: {datetime.now().strftime('%d/%m/%Y %H:%M')}", st["sub"]),
            Spacer(1, 10),
            HRFlowable(width="100%", thickness=1, color=colors.HexColor("#d1d5db"), spaceAfter=10)
        ]

        table_data = [[
            Paragraph("<b>No</b>", st["th_center"]),
            Paragraph("<b>ID Nasabah</b>", st["th_center"]),
            Paragraph("<b>NIK</b>", st["th_center"]),
            Paragraph("<b>Nama Lengkap</b>", st["th_left"]),
            Paragraph("<b>Kategori</b>", st["th_center"]),
            Paragraph("<b>No. HP</b>", st["th_center"]),
            Paragraph("<b>Status</b>", st["th_center"]),
            Paragraph("<b>Saldo (Rp)</b>", st["th_right"])
        ]]

        total_bal = 0
        for i, n in enumerate(nasabahs, 1):
            bal = n.get("balance", 0)
            total_bal += bal
            table_data.append([
                Paragraph(str(i), st["cell_center"]),
                Paragraph(n["customer_id"], st["cell_center"]),
                Paragraph(n["nik"], st["cell_center"]),
                Paragraph(n["name"], st["cell_left"]),
                Paragraph(n.get("nasabah_category") or "Rumah Tangga/Individu", st["cell_center"]),
                Paragraph(n["phone"], st["cell_center"]),
                Paragraph(n["status"], st["cell_center"]),
                Paragraph(format_rupiah(bal), st["cell_right"]),
            ])

        table_data.append([
            Paragraph("<b>TOTAL SALDO NASABAH</b>", st["cell_bold_center"]),
            "", "", "", "", "", "",
            Paragraph(f"<b>{format_rupiah(total_bal)}</b>", st["cell_bold_right"]),
        ])

        t = Table(table_data, colWidths=[25, 65, 85, 110, 75, 75, 42, 70])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f0fdf4")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#166534")),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('SPAN', (0, -1), (6, -1)),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#f9fafb")),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(t)
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    # 4. PRICE MASTER REPORT
    @classmethod
    def export_prices_excel(cls) -> bytes:
        query = """
            SELECT p.*, c.name as category_name
            FROM waste_price_masters p
            JOIN waste_categories c ON p.category_id = c.id
            ORDER BY c.name ASC, p.status ASC, p.effective_date DESC
        """
        prices = db.fetch_all(query)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daftar Harga"

        cls._create_excel_header(ws, "LAPORAN DAFTAR HARGA SAMPAH", f"Per Tanggal: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 7)

        headers = ["No", "Kelompok Sampah", "Kode Harga", "Harga per Kg (Rp)", "Tanggal Efektif", "Status", "Catatan"]
        header_row = 5
        header_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        header_font = Font(name="Arial", size=9, bold=True, color="166534")
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'), right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'), bottom=Side(style='thin', color='E5E7EB')
        )

        for col_num, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        current_row = 6
        for i, p in enumerate(prices, 1):
            ws.cell(row=current_row, column=1, value=i).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=current_row, column=2, value=p["category_name"]).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.cell(row=current_row, column=3, value=p.get("price_code") or "-").alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=current_row, column=4, value=p["price_per_kg"]).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row=current_row, column=4).number_format = "#,##0"
            ws.cell(row=current_row, column=5, value=format_date(p["effective_date"])).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=current_row, column=6, value=p["status"]).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=current_row, column=7, value=p.get("notes") or "-").alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            for c in range(1, 8):
                ws.cell(row=current_row, column=c).border = thin_border
                ws.cell(row=current_row, column=c).font = Font(name="Arial", size=9)
            current_row += 1

        cls._auto_fit_columns(ws, 1, 7)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @classmethod
    def export_prices_pdf(cls) -> bytes:
        query = """
            SELECT p.*, c.name as category_name
            FROM waste_price_masters p
            JOIN waste_categories c ON p.category_id = c.id
            ORDER BY c.name ASC, p.status ASC, p.effective_date DESC
        """
        prices = db.fetch_all(query)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=portrait(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)

        st = cls._get_pdf_styles()

        elements = [
            Paragraph(settings.BANK_NAME.upper(), st["title"]),
            Paragraph("LAPORAN DAFTAR HARGA SAMPAH", st["title"]),
            Paragraph(f"Per Tanggal: {datetime.now().strftime('%d/%m/%Y %H:%M')}", st["sub"]),
            Spacer(1, 10),
            HRFlowable(width="100%", thickness=1, color=colors.HexColor("#d1d5db"), spaceAfter=10)
        ]

        table_data = [[
            Paragraph("<b>No</b>", st["th_center"]),
            Paragraph("<b>Kelompok Sampah</b>", st["th_left"]),
            Paragraph("<b>Kode Harga</b>", st["th_center"]),
            Paragraph("<b>Harga/kg (Rp)</b>", st["th_right"]),
            Paragraph("<b>Tgl Efektif</b>", st["th_center"]),
            Paragraph("<b>Status</b>", st["th_center"]),
            Paragraph("<b>Catatan</b>", st["th_left"])
        ]]

        for i, p in enumerate(prices, 1):
            table_data.append([
                Paragraph(str(i), st["cell_center"]),
                Paragraph(p["category_name"], st["cell_left"]),
                Paragraph(p.get("price_code") or "-", st["cell_center"]),
                Paragraph(format_rupiah(p["price_per_kg"]), st["cell_right"]),
                Paragraph(format_date(p["effective_date"]), st["cell_center"]),
                Paragraph(p["status"], st["cell_center"]),
                Paragraph(p.get("notes") or "-", st["cell_left"]),
            ])

        t = Table(table_data, colWidths=[28, 145, 65, 80, 75, 54, 100])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f0fdf4")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#166534")),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(t)
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()


report_service = ReportService()
